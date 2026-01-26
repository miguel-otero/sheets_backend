import io
import os
import re
import time
import tempfile
import logging
from typing import List, Optional, Literal

from fastapi import FastAPI, HTTPException
from pydantic import BaseModel, Field

import google.auth
import httplib2
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaIoBaseDownload
from google_auth_httplib2 import AuthorizedHttp

from openpyxl import load_workbook

logger = logging.getLogger("sheets_automation")
logging.basicConfig(level=logging.INFO)

# ------------------------------------------------------------
# Config
# ------------------------------------------------------------
DRIVE_SHEETS_SCOPES = [
    "https://www.googleapis.com/auth/drive",
    "https://www.googleapis.com/auth/spreadsheets",
]


# Ajusta si quieres. 60s suele funcionar bien para llamadas largas.
HTTP_TIMEOUT_SECONDS = int(os.getenv("GOOGLE_HTTP_TIMEOUT", "60"))


# ------------------------------------------------------------
# Models
# ------------------------------------------------------------
class ConvertXlsxToExistingSheetsRequest(BaseModel):
    xlsx_file_id: str = Field(..., description="Drive fileId o URL del XLSX")
    target_spreadsheet_id: str = Field(..., description="SpreadsheetId o URL del Google Sheets existente (destino)")
    selected_tabs: List[str] = Field(..., min_items=1, description="Pestañas del XLSX a copiar")

    wipe_mode: Literal["all", "selected"] = Field(
        "all",
        description="all=borrar valores de TODAS las hojas del destino. selected=borrar solo hojas seleccionadas."
    )

    value_input_option: Literal["RAW", "USER_ENTERED"] = Field("RAW")
    target_cells_per_request: int = Field(80000, ge=10000, le=200000)
    max_retries: int = Field(8, ge=1, le=15)
    xlsx_file_id: str = Field(..., description="Drive fileId o URL del XLSX")
    selected_tabs: List[str] = Field(..., min_items=1, description="Pestañas del XLSX a copiar (ej: ['mov_general'])")

    destination_folder_id: Optional[str] = Field(
        None, description="Drive folderId o URL destino. Si no se envía, usa la carpeta padre del XLSX."
    )
    new_spreadsheet_name: Optional[str] = Field(
        None, description="Nombre del Google Sheets destino. Si no se envía, usa '<xlsx_name> (seleccion)'."
    )

    value_input_option: Literal["RAW", "USER_ENTERED"] = Field(
        "RAW",
        description="RAW es más rápido/robusto. USER_ENTERED interpreta fórmulas/formatos tipo Sheets.",
    )

    # Control de chunking
    target_cells_per_request: int = Field(
        80000, ge=10000, le=200000, description="Aprox celdas por request a Sheets (ajusta si da error por tamaño)."
    )
    max_retries: int = Field(8, ge=1, le=15, description="Reintentos ante 429/5xx")


# ------------------------------------------------------------
# Helpers (Drive/Sheets)
# ------------------------------------------------------------

def extract_drive_id(s: str) -> str:
    m = re.search(r"[-\w]{25,}", s or "")
    return m.group(0) if m else s

def build_google_services():
    creds, _ = google.auth.default(scopes=DRIVE_SHEETS_SCOPES)
    http = httplib2.Http(timeout=HTTP_TIMEOUT_SECONDS)
    authed_http = AuthorizedHttp(creds, http=http)
    drive_service = build("drive", "v3", http=authed_http, cache_discovery=False)
    sheets_service = build("sheets", "v4", http=authed_http, cache_discovery=False)
    return drive_service, sheets_service

def with_retries(fn, *, max_retries=8, base_sleep=1.0, retryable_status=(429, 500, 502, 503, 504)):
    for attempt in range(max_retries):
        try:
            return fn()
        except HttpError as e:
            status = getattr(e.resp, "status", None)
            if status in retryable_status and attempt < max_retries - 1:
                time.sleep(base_sleep * (2**attempt) + (0.1 * attempt))
                continue
            raise

def get_drive_file_metadata(drive_service, file_id: str, fields: str = "id,name,parents,mimeType,size"):
    return with_retries(lambda: drive_service.files().get(
        fileId=file_id, fields=fields, supportsAllDrives=True
    ).execute())

def download_drive_file(drive_service, file_id: str, out_path: str, *, max_retries: int, chunk_size: int = 8 * 1024 * 1024):
    request = drive_service.files().get_media(fileId=file_id, supportsAllDrives=True)
    fh = io.FileIO(out_path, "wb")
    downloader = MediaIoBaseDownload(fh, request, chunksize=chunk_size)
    done = False
    while not done:
        status, done = with_retries(lambda: downloader.next_chunk(), max_retries=max_retries)
        if status:
            pct = int(status.progress() * 100)
            if pct % 10 == 0:
                logger.info("Descargando XLSX... %s%%", pct)
    fh.close()

def col_to_a1_letter(n: int) -> str:
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s

def a1_start(sheet_title: str, start_row: int, start_col: int = 1) -> str:
    return f"'{sheet_title}'!{col_to_a1_letter(start_col)}{start_row}"

def choose_chunk_rows(num_cols: int, target_cells_per_request: int, min_rows: int = 200, max_rows: int = 5000) -> int:
    if num_cols <= 0:
        return min_rows
    rows = target_cells_per_request // max(1, num_cols)
    return int(max(min_rows, min(max_rows, rows)))

def safe_cell(v):
    if v is None:
        return ""
    try:
        import datetime as dt
        if isinstance(v, (dt.datetime, dt.date)):
            return v.isoformat()
    except Exception:
        pass
    return v

def validate_sheet_title(title: str) -> str:
    forbidden = set(r'[]:*?/\\')
    if any(ch in forbidden for ch in title):
        raise ValueError(f"Nombre de hoja inválido para Google Sheets: '{title}'")
    if len(title) > 100:
        raise ValueError(f"Nombre de hoja demasiado largo (>100): '{title}'")
    if title.strip() == "":
        raise ValueError("Nombre de hoja vacío no permitido.")
    return title

def _http_error_text(e: HttpError) -> str:
    content = getattr(e, "content", b"") or b""
    try:
        return content.decode("utf-8", errors="ignore")
    except Exception:
        return str(content)

def is_payload_too_large(e: HttpError) -> bool:
    status = getattr(e.resp, "status", None)
    txt = _http_error_text(e).lower()
    return (
        status in (413,) or
        "request payload size exceeds" in txt or
        "payload exceeds limit" in txt or
        "entity too large" in txt
    )

def flush_values_resilient(
    sheets_service,
    spreadsheet_id: str,
    sheet_title: str,
    start_row: int,
    values_chunk,
    *,
    value_input_option: str,
    max_retries: int,
):
    """Escribe un chunk; si el payload es grande, divide y reintenta."""
    if not values_chunk:
        return

    rng = a1_start(sheet_title, start_row, 1)
    end_row = start_row + len(values_chunk) - 1

    try:
        logger.info("Escribiendo %s! filas %s-%s (%s filas)", sheet_title, start_row, end_row, len(values_chunk))
        with_retries(lambda: sheets_service.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id,
            range=rng,
            valueInputOption=value_input_option,
            body={"values": values_chunk},
        ).execute(), max_retries=max_retries)

    except HttpError as e:
        # LOG DEL ERROR REAL
        logger.warning("Fallo escribiendo %s filas %s-%s. status=%s body=%s",
                       sheet_title, start_row, end_row, getattr(e.resp, "status", None), _http_error_text(e)[:1000])

        # AUTOSPLIT si es payload grande
        if is_payload_too_large(e) and len(values_chunk) > 1:
            mid = len(values_chunk) // 2
            flush_values_resilient(sheets_service, spreadsheet_id, sheet_title, start_row, values_chunk[:mid],
                                   value_input_option=value_input_option, max_retries=max_retries)
            flush_values_resilient(sheets_service, spreadsheet_id, sheet_title, start_row + mid, values_chunk[mid:],
                                   value_input_option=value_input_option, max_retries=max_retries)
            return
        raise

def resize_sheets_grid(sheets_service, spreadsheet_id: str, title_to_id: dict, title_to_size: dict, *, max_retries: int):
    """
    title_to_size: { "Hoja": (rowCount, columnCount) }
    """
    requests = []
    for title, (r, c) in title_to_size.items():
        sheet_id = title_to_id[title]
        requests.append({
            "updateSheetProperties": {
                "properties": {
                    "sheetId": sheet_id,
                    "gridProperties": {
                        "rowCount": int(max(1, r)),
                        "columnCount": int(max(1, c)),
                    },
                },
                "fields": "gridProperties.rowCount,gridProperties.columnCount",
            }
        })

    if requests:
        with_retries(lambda: sheets_service.spreadsheets().batchUpdate(
            spreadsheetId=spreadsheet_id,
            body={"requests": requests}
        ).execute(), max_retries=max_retries)

# ------------------------------------------------------------
# Core conversion
# ------------------------------------------------------------

def write_xlsx_tabs_into_existing_spreadsheet(
    *,
    drive_service,
    sheets_service,
    xlsx_file_id: str,
    target_spreadsheet_id: str,
    selected_tabs: List[str],
    wipe_mode: str,
    value_input_option: str,
    target_cells_per_request: int,
    max_retries: int,
):
    xlsx_file_id = extract_drive_id(xlsx_file_id)
    target_spreadsheet_id = extract_drive_id(target_spreadsheet_id)
    selected_tabs = [validate_sheet_title(t) for t in selected_tabs]

    # 1) Verificar acceso al spreadsheet destino + obtener hojas existentes
    try:
        ss_meta = with_retries(lambda: sheets_service.spreadsheets().get(
            spreadsheetId=target_spreadsheet_id,
            fields="properties(title),sheets(properties(sheetId,title,gridProperties(rowCount,columnCount)))"
        ).execute(), max_retries=max_retries)

        title_to_id = {s["properties"]["title"]: s["properties"]["sheetId"] for s in ss_meta.get("sheets", [])}
        existing_titles = list(title_to_id.keys())
    except HttpError as e:
        raise ValueError("No tengo acceso al Google Sheets destino (comparte el archivo con la service account como Editor).") from e

    dest_title = ss_meta["properties"]["title"]
    existing_sheets = ss_meta.get("sheets", [])
    existing_titles = [s["properties"]["title"] for s in existing_sheets]

    # 2) Descargar XLSX
    tmp = tempfile.NamedTemporaryFile(prefix="xlsx_", suffix=".xlsx", dir="/tmp", delete=False)
    tmp_path = tmp.name
    tmp.close()

    logger.info("Descargando XLSX (fileId=%s) a %s", xlsx_file_id, tmp_path)

    try:
        download_drive_file(drive_service, xlsx_file_id, tmp_path, max_retries=max_retries)
        logger.info("Descarga completada: %s", tmp_path)

        # 3) Cargar XLSX streaming
        wb = load_workbook(filename=tmp_path, read_only=True, data_only=True)
        available = wb.sheetnames
        missing = [t for t in selected_tabs if t not in available]
        if missing:
            raise ValueError(f"Pestañas no encontradas en el XLSX: {missing}. Disponibles: {available}")

        # 4) Asegurar que existen las hojas destino (crear si faltan)
        to_add = [t for t in selected_tabs if t not in existing_titles]
        if to_add:
            add_requests = [{"addSheet": {"properties": {"title": t}}} for t in to_add]
            with_retries(lambda: sheets_service.spreadsheets().batchUpdate(
                spreadsheetId=target_spreadsheet_id,
                body={"requests": add_requests}
            ).execute(), max_retries=max_retries)

            # refrescar títulos existentes
            ss_meta = with_retries(lambda: sheets_service.spreadsheets().get(
                spreadsheetId=target_spreadsheet_id,
                fields="sheets(properties(title))"
            ).execute(), max_retries=max_retries)
            existing_titles = [s["properties"]["title"] for s in ss_meta.get("sheets", [])]

        # 5) BORRAR CONTENIDO (wipe)
        # Desinfla todo (si wipe_mode="all") para liberar celdas del workbook
        if wipe_mode == "all":
            shrink_sizes = {t: (1, 1) for t in existing_titles}
            resize_sheets_grid(sheets_service, target_spreadsheet_id, title_to_id, shrink_sizes, max_retries=max_retries)

        # Ajusta las hojas destino al tamaño que vas a necesitar (según max_row/max_col del XLSX)
        target_sizes = {}
        for tab in selected_tabs:
            ws = wb[tab]
            target_sizes[tab] = (ws.max_row or 1, ws.max_column or 1)

        resize_sheets_grid(sheets_service, target_spreadsheet_id, title_to_id, target_sizes, max_retries=max_retries)


        logger.info("Wipe completado (%s).", wipe_mode)

        # 6) Escribir hojas seleccionadas en chunks
        def flush_values_resilient(sheet_title: str, start_row: int, values_chunk):
            if not values_chunk:
                return
            rng = a1_start(sheet_title, start_row, 1)
            with_retries(lambda: sheets_service.spreadsheets().values().update(
                spreadsheetId=target_spreadsheet_id,
                range=rng,
                valueInputOption=value_input_option,
                body={"values": values_chunk}
            ).execute(), max_retries=max_retries)

        for tab in selected_tabs:
            ws = wb[tab]
            max_col = ws.max_column or 0
            chunk_rows = choose_chunk_rows(max_col, target_cells_per_request=target_cells_per_request)

            # tope duro para evitar requests gigantes
            chunk_rows = min(chunk_rows, 1000)

            logger.info("Copiando hoja '%s' (cols≈%s) en chunks de %s filas", tab, max_col, chunk_rows)

            row_cursor = 1
            buffer = []
            for row in ws.iter_rows(values_only=True):
                buffer.append([safe_cell(v) for v in row])
                if len(buffer) >= chunk_rows:
                    flush_values_resilient(tab, row_cursor, buffer)
                    row_cursor += len(buffer)
                    buffer = []

            if buffer:
                flush_values_resilient(tab, row_cursor, buffer)

            logger.info("Hoja '%s' copiada.", tab)

        return {
            "spreadsheet_id": target_spreadsheet_id,
            "spreadsheet_url": f"https://docs.google.com/spreadsheets/d/{target_spreadsheet_id}",
            "name": dest_title,
            "wipe_mode": wipe_mode,
        }

    finally:
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        except Exception:
            logger.warning("No pude borrar temp file: %s", tmp_path)

# Debug

def drive_client():
    creds, _ = google.auth.default(scopes=SCOPES)
    http = httplib2.Http(timeout=60)
    authed = AuthorizedHttp(creds, http=http)
    return build("drive", "v3", http=authed, cache_discovery=False)
